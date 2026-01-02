import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from collections import defaultdict
import random

# -------------------- منطق اصلی --------------------
def process_files(exam_path, monitor_path, student_limit):

    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    exams = pd.read_excel(exam_path)
    monitors = pd.read_excel(monitor_path)

    monitors["count"] = 0
    main_monitors = monitors.iloc[:5].copy()
    other_monitors = monitors.iloc[5:].copy()

    all_days = sorted(exams["روز امتحان"].unique())
    all_results = []

    for day in all_days:

        day_exams = exams[exams["روز امتحان"] == day].copy()
        day_exams["ناظر رسمی"] = ""
        day_exams["ناظر سرکشی"] = ""
        day_exams["reason_official"] = ""  # علت ناظر رسمی

        # تشخیص پنج‌شنبه
        is_thursday = day_exams["روز هفته"].astype(str).str.contains("پنج").any()

        # انتخاب ناظرهای اصلی
        main_candidates = main_monitors.copy()
        if is_thursday:
            main_candidates = main_candidates[main_candidates["جنسیت"] == 2]
        if len(main_candidates) < 3:
            raise Exception(f"ناظر اصلی کافی برای روز {day} وجود ندارد")

        patrols = main_candidates.sample(2)
        patrol_names = patrols["نام ناظر"].tolist()
        main_officials = main_candidates[~main_candidates["نام ناظر"].isin(patrol_names)]

        # نیاز به ناظر رسمی
        day_exams["need_official"] = False
        day_exams.loc[day_exams["تعداد"] >= student_limit, "need_official"] = True

        # همزمانی استاد + محل امتحان
        grp = day_exams.groupby(["نام استاد", "ساعت شروع"])
        for (_, _), g in grp:
            if len(g) > 1:
                if g["محل امتحان"].nunique() == 1:
                    total_students = g["تعداد"].sum()
                    if total_students < student_limit:
                        continue
                need = len(g) - 1
                idxs = g.sort_values("تعداد", ascending=False).index[:need]
                day_exams.loc[idxs, "need_official"] = True

        # تخصیص ناظر رسمی
        used_today = set()
        
        pool = pd.concat([main_officials, other_monitors]).sort_values("count")
        pool_main = pd.concat([main_officials , ""]).sort_values("count")       #HHHHH

        pool_other = pd.concat([other_monitors , ""]).sort_values("count")      #HHHH
	
        for idx, row in day_exams.iterrows():
            if not row["need_official"]:
                day_exams.at[idx, "reason_official"] = "ناظر رسمی نمی‌خواهد"
                continue

            eligible = pool.copy()
            if is_thursday:
                eligible = eligible[eligible["جنسیت"] == 2]
            if str(row["ساعت شروع"]) == "14":
                eligible = eligible[eligible["جنسیت"] == 2]

            eligible = eligible[~eligible["نام ناظر"].isin(used_today)]
            if eligible.empty:
                day_exams.at[idx, "ناظر رسمی"] = "ناظر کافی موجود نیست"
                day_exams.at[idx, "reason_official"] = "کمبود ناظر"
                continue

            chosen = eligible.iloc[0]
            name = chosen["نام ناظر"]
            day_exams.at[idx, "ناظر رسمی"] = name
            used_today.add(name)
            monitors.loc[monitors["نام ناظر"] == name, "count"] += 1
            pool.loc[pool["نام ناظر"] == name, "count"] += 1

            if row["تعداد"] >= student_limit:
                day_exams.at[idx, "reason_official"] = "تعداد دانشجو زیاد"
            else:
                day_exams.at[idx, "reason_official"] = "همزمانی استاد"

        # ناظر سرکشی
        no_official = day_exams[day_exams["ناظر رسمی"] == ""]
        patrol_cycle = patrol_names * 100
        for i, idx in enumerate(no_official.index):
            day_exams.at[idx, "ناظر سرکشی"] = patrol_cycle[i]

        all_results.append(day_exams)

    final_df = pd.concat(all_results)

    # ---------- خروجی اکسل با رنگ و جمع روزانه ----------
    with pd.ExcelWriter("all_exams_with_monitors.xlsx", engine="openpyxl") as writer:
        for day in all_days:
            df_day = final_df[final_df["روز امتحان"] == day].copy()
            df_day.to_excel(writer, sheet_name=f"روز {day}", index=False)

        writer.save()

    # باز کردن برای فرمت دهی رنگ
    wb = load_workbook("all_exams_with_monitors.xlsx")
    for day in all_days:
        ws = wb[f"روز {day}"]
        max_row = ws.max_row
        max_col = ws.max_column

        reason_col = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == "reason_official":
                reason_col = col_idx
                break
        if not reason_col:
            continue

        # رنگ‌بندی سلول‌ها
        for row in range(2, max_row + 1):
            reason = ws.cell(row=row, column=reason_col).value
            for col in range(1, max_col + 1):
                if reason == "ناظر رسمی نمی‌خواهد":
                    ws.cell(row=row, column=col).font = Font(color="0000FF")  # آبی
                elif reason == "تعداد دانشجو زیاد":
                    ws.cell(row=row, column=col).font = Font(color="008000")  # سبز
                elif reason == "همزمانی استاد":
                    ws.cell(row=row, column=col).font = Font(color="FFA500")  # نارنجی
                elif reason == "کمبود ناظر":
                    ws.cell(row=row, column=col).font = Font(color="FF0000")  # قرمز

        # ---------- جمع روزانه ----------
        cnt_students = sum(1 for row in range(2, max_row + 1)
                           if ws.cell(row=row, column=reason_col).value == "تعداد دانشجو زیاد")
        cnt_simult = sum(1 for row in range(2, max_row + 1)
                         if ws.cell(row=row, column=reason_col).value == "همزمانی استاد")
        cnt_missing = sum(1 for row in range(2, max_row + 1)
                          if ws.cell(row=row, column=reason_col).value == "کمبود ناظر")

        ws.append([])
        ws.append(["جمع روزانه"])
        ws.append(["ناظر به دلیل تعداد دانشجو", cnt_students])
        ws.append(["ناظر به دلیل همزمانی استاد", cnt_simult])
        ws.append(["امتحان بدون ناظر کافی", cnt_missing])

    wb.save("all_exams_with_monitors.xlsx")

    # ---------- گزارش ناظر رسمی ----------
    officials = final_df[final_df["ناظر رسمی"].notna() & (final_df["ناظر رسمی"] != "") & (final_df["ناظر رسمی"] != "ناظر کافی موجود نیست")]
    with pd.ExcelWriter("official_monitor_detailed_report.xlsx", engine="openpyxl") as writer:
        for name, g in officials.groupby("ناظر رسمی"):
            g = g.sort_values(["روز امتحان", "ساعت شروع"])
            g.to_excel(writer, sheet_name=name[:31], index=False)

    print("✅ شیت‌های روزانه رنگی و جمع روزانه اضافه شد")


# -------------------- رابط گرافیکی --------------------

def browse_exam():
    path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    exam_var.set(path)

def browse_monitor():
    path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    monitor_var.set(path)

def run():
    try:
        exam_path = exam_var.get()
        monitor_path = monitor_var.get()
        limit = int(limit_var.get())

        if not exam_path or not monitor_path:
            messagebox.showerror("خطا", "لطفاً هر دو فایل را انتخاب کنید")
            return

        process_files(exam_path, monitor_path, limit)
        messagebox.showinfo("موفق", "پردازش با موفقیت انجام شد")

    except Exception as e:
        messagebox.showerror("خطا", str(e))


# -------------------- ساخت فرم --------------------

root = tk.Tk()
root.title("سیستم تخصیص ناظر امتحانات")
root.geometry("620x260")

exam_var = tk.StringVar()
monitor_var = tk.StringVar()
limit_var = tk.StringVar(value="30")

tk.Label(root, text="فایل امتحانات:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
tk.Entry(root, textvariable=exam_var, width=55).grid(row=0, column=1)
tk.Button(root, text="انتخاب", command=browse_exam).grid(row=0, column=2)

tk.Label(root, text="فایل ناظرها:").grid(row=1, column=0, padx=10, pady=10, sticky="w")
tk.Entry(root, textvariable=monitor_var, width=55).grid(row=1, column=1)
tk.Button(root, text="انتخاب", command=browse_monitor).grid(row=1, column=2)

tk.Label(root, text="حد آستانه تعداد دانشجو:").grid(row=2, column=0, padx=10, pady=10, sticky="w")
tk.Entry(root, textvariable=limit_var, width=10).grid(row=2, column=1, sticky="w")

tk.Button(
    root,
    text="اجرای تخصیص ناظرها",
    command=run,
    bg="#2c7be5",
    fg="white",
    height=2,
    width=30
).grid(row=4, column=1, pady=30)

root.mainloop()
